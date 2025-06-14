from openpyxl import load_workbook

#1 - Estanciando o load_workbook
wb = load_workbook(filename='files/teste.xlsx') #Informando o caminho onde está o caminho
planilha1 = wb['Planilha 1'] #Identificando que planilha1 = se refere a folha dentro do nosso workbook

#2 - Acessando valores
print(planilha1['B2'].value) #Acessando valor de uma determinada posição da planilha

#3 - Iterando para acessar todos os valores
#Forma 1:
print("---------------------------------------------")
for i in range(1, 5):
    # Dentro dos parametros vemos que %s é o valor a ser preenchido e %i é o valor de ir que ira atribuir %s para ser A1, A2 e A3
    print(f"{planilha1[f'A%s' %i].value} {planilha1[f'B%s' %i].value} {planilha1[f'C%s' %i].value}")

#forma 2:
print("---------------------------------------------")
for i in range(2,5):
    ano = planilha1[f'A%s' %i].value
    lucro = planilha1[f'B%s' %i].value
    custo = planilha1[f'C%s' %i].value
    print("{0} teve {1} de lucro e {2} de custo".format(ano, lucro, custo))