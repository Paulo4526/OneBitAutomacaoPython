#Realizaremos operações com planilhas via python
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

#1 - Inicializando o load_workbook
wb = load_workbook(filename='files/gastos.xlsx') #Aqui passaremos o diretório absoluto com o nome da planilha que queremos ler
planilha = wb['Sheet'] #Defininado qual planilha utilizaremos
print(planilha)

#Realizando as operações
valor_total = 0
#2 - Operações de soma
for i in range(2, planilha.max_row): #Utilizamos o max_row para sabermos quantas linhas preenchidas temos ness aplanilha
    if planilha['B%s' %i].value is None or int(planilha['B%s' %i].value) == int:
        continue
    else:
        print(f"Tipo de Gasto: {planilha['A%s' % i].value}, Valor R${planilha['B%s' % i].value}")
        valor = int(planilha['B%s' % i].value)
        valor_total += valor
        if i == planilha.max_row - 2:
            # Não é preciso colocar essa lógica if dentro da operação, pois a alteração está endo feita na variavel global valor_total
            # esse valor total pode ser printado após a iteração fora do laço de repetição
            print(f"Valor total de gatos: {valor_total}")
            planilha['A10'] = 'Total'  # Alterando na planilha
            planilha['B10'] = valor_total  # Alterando na planilha
            wb.save(filename='files/gastos.xlsx') #Salvando as alterações

#3 - Mesclando células
planilha['A9'] = 'Teste'
planilha.merge_cells('A9:B9') #Mesclando
planilha.unmerge_cells('A9:B9') #Desfazendo a mesclagem
#Para ver a merclagem é necessário abrir a planilha via excel ou algum aplicativo fora da IDE
# wb.save(filename='files/gastos.xlsx') #Salvando as alterações

#3 - Inseringo imagens
img = Image('files/bb.preco.png')
planilha.add_image(img, 'C9')
# Para ver a imagem é necessário abrir a planilha via excel ou algum aplicativo fora da IDE
# wb.save(filename='files/gastos.xlsx') #Salvando as alterações

#4 - Deletando Células
planilha.delete_rows(2) #Deletando a penultima linha da planilha
planilha.delete_cols(3) #Deletando a ultima coluna da planilha
wb.save(filename='files/gastos2.xlsx') #Salvando os dados em um outro arquivo


