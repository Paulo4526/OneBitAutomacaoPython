from openpyxl import load_workbook, workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.workbook import Workbook

dict_ano = {}

#1 - Importanto Despesas
arquivo1 = load_workbook(filename='files/Despesas.xlsx')
ws1 = arquivo1['despesas'] #Recuperando valores da planilha despesas
max_linhas = ws1.max_row

#Recuperando valores das despesas e salvando no dicionario
for i in range(2, max_linhas + 1): #Recuperando os valores a partir da linha 2, não recebendo os valores dos títulos na linha 1
    print(f"Ano: {ws1['A%s' %i].value} Valor: {ws1['B%s' %i].value}") #Print para mostrar o valor da iteração
    dict_ano[ws1['A%s' %i].value] = {'despesa': ws1['B%s' %i].value, 'receita': 0}

#2 - Importanto Receita
arquivo2 = load_workbook(filename='files/Receitas.xlsx')
ws2 = arquivo2['receita']
max_linhas2 = ws2.max_row

for i in range(2, max_linhas2 + 1):
    print(f"Ano: {ws2['A%s' %i].value} Valor: {ws2['B%s' %i].value}")
    dict_ano[ws2['A%s' %i].value]['receita'] = ws2['B%s' %i].value

print(dict_ano)

#3 - Criando a planilha
wb = Workbook()
ws = wb.active

ws['A1'] = 'Ano'
ws['B1'] = 'Despesas'
ws['C1'] = 'Receitas'

for idx, (key, value) in enumerate(dict_ano.items(), start=2):
    ws[f'A{idx}'] = key
    ws[f'B{idx}'] = value['despesa']
    ws[f'C{idx}'] = value['receita']

#4 - Criando o grafico em Barras
chart1 = BarChart()
chart1.type = 'col'
chart1.style = 12
chart1.title = 'Receita x Despesa por Ano'
chart1.x_axis.title = 'Ano'
chart1.y_axis.title = 'R$'

category = Reference(
    ws,
    min_col= 1,
    max_col= 1,
    min_row= 2,
    max_row= ws.max_row
)

dados = Reference(
    ws,
    min_col= 2,
    max_col= 3,
    min_row= 1,
    max_row=ws.max_row
)

chart1.add_data(dados, titles_from_data=True)
chart1.set_categories(category)
chart1.shape = 4
ws.add_chart(chart1, 'A%s' %(ws.max_row + 2))

wb.save(filename='files/despesa-unida.xlsx')