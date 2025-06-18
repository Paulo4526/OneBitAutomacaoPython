from openpyxl import Workbook
from openpyxl.chart import AreaChart, Reference

#1 - Instanciando o WorkBook
wb = Workbook()
#Iniciando o WorkBook
ws = wb.active
#Criando um array que irá preencher uma planilha e após isso criaremos um grafico baseado nessa planilha.
data = [
    ['Ano', 'Lucro', 'Custos'],
    [2017, 40, 30],
    [2018, 35, 25],
    [2019, 30, 20],
    [2020, 10, 10],
    [2021, 15, 10],
    [2022, 25, 15]
]


#2 - Criando a panilha
for d in data:
    ws.append(d)


#3 - Criando o grafico
#Estanciando o AreaChart
chart = AreaChart()
chart.title = 'Lucro x Custos por Ano'
chart.style = 13
chart.x_axis.title = 'Ano' #Aqui definimos para o eixo x, o título para sua funcionalidade que será mostrar o ano
chart.y_axis.title = 'Gastos e Lucros em %' #Aqui definimos para o eixo y, o título para sua funcionalidade que será mostrar os gastos e os lucros em %

#Aqui criaremos a referência do eixo X contendo os anos
categorias = Reference( #Criamos uma variavel categorias para referenciar quais os tópicos que serão abordadps
    ws, #Informando qual a nossa planilha no WS
    min_col=1, #Aqui informamos qual coluna iremos pegar iniciando da coluna 1
    max_col=1, #Aqui informamos em qual coluna iremos terminar, sendo a própria coluna 1 onde será informado a categoria
    min_row=2, #Aqui informamos a partir de qual linha iremos iniciar a referencia
    max_row= ws.max_row #Aqui informamos até qual linha iremos referenciar nossa categoria
    #OBS: No grafico criado, vemos que abaixo do gráfico teremos Somente os anos como referência da categoria. Caso coloquemos o min_row como 1, irá aparecer o ano abaixo no gráfico
)

#Definindo os valores de referência para o eixo y
dados = Reference(
    ws, #Informando de qual planilha virá os dados
    min_col= 2, #Informando que as informações serão pegos a partir da 2 coluna da planilha
    min_row= 1, #Informando que os valores serão obtidas desde a linha 1, sendo a primeira o nome da série
    max_col= 3, #Informando que as informações serão resgatadas somente até a coluna 3
    max_row= ws.max_row #Informando que os valores serão obtidos até o máximo de linhas existentes na planilha
)

chart.add_data(dados, titles_from_data=True) #Aqui definimos os parametros para o eixo y, e o título true é somente para mostrar em qual categoria a séria se encaixa, como lucro ou custos
chart.set_categories(categorias) #Aqui definimos os parametros para o eixo x
ws.add_chart(chart, "A10") #Adicionando o grafico na coluna A linha 10 da planilha

wb.save(filename='files/novo-grafico.xlsx')

#Explicação simplificada do código para maiores esclarecimentos:
#Variavel Dados para o eixo Y
#Ao definirmos o min_row para 1, iniciando da coluna 2 sendo nossa planilha da seguinte forma:
'''
Lucro Custos -> Ao definirmos o min_row será mostrado os títulos na série como Lucro para a coluna B e Custos para C
  40    30   -> Caso definirmos o min_row para 2, será mostrado na série os valores 40 e 30, pois a primeira linha é sempre para informar a Série de medição. 
  35    25
  30    20
'''