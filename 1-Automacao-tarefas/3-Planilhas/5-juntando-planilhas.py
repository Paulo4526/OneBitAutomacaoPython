#Juntando as planilhas em diferentes arquivos para um único arquivo
from openpyxl import Workbook, load_workbook

lista_arquivos = ['gastos', 'gastos2', 'teste'] #aqui é uma lista com nome tanto de arquivos quanto nome de suas planilhas que serão iteradas
#1 - Inicianlizando meu Workbook
wb = Workbook()
nome_arquivo = 'files/resultado.xlsx'

for nome in lista_arquivos:
    arquivo = load_workbook(filename='files/%s.xlsx' %nome) #Realizará a leitura de cada arquivo
    #Como iremos criar um novo arquivo com 3 planilhas, a cada iteração, será lido o arquivo e será criado uma nova planilha e inserido esses valores na nova planilha
    sheet = arquivo[nome]
    max_lines = sheet.max_row #Recuperará o máximo de linhas que contem no arquivo
    max_columns = sheet.max_column #Recuperará o máximo de colunas que tem no arquivo
    ws = wb.create_sheet(title=nome) #Criará uma planilha no arquivo com o nome do arquivo a ser transferido
    #Iterando valores da planilha
    for line in range(1, max_lines + 1): #Iterando a linha
        for col in range(1, max_columns + 1): #Iterando a coluna
            data = sheet.cell(row=line, column=col)
            ws.cell(row=line, column=col).value = data.value

wb.remove(wb['Sheet'])
wb.save(nome_arquivo)